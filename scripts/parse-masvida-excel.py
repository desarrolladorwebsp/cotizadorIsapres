#!/usr/bin/env python3
"""Parsea los Excel de Nueva Masvida en storage/planes-pdf/mas vida / a JSON."""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

ISAPRE_NAME = "Nueva Masvida"
DEFAULT_FOLDER = Path(__file__).resolve().parent.parent / "storage" / "planes-pdf" / "mas vida "
OUTPUT_PATH = Path(__file__).resolve().parent / ".masvida-plans-parsed.json"

PLAN_CODE_HEADERS = {"PLAN", "CODIGO PLAN", "CÓDIGO PLAN"}
PRICE_HEADERS = {"PRECIO BASE"}
HOSPITAL_HEADERS = {"HOSPITALARIO", "PORCENTAJE HOSPITALARIO", "HOSPITALARIO "}
AMBULATORY_HEADERS = {"AMBULATORIO", "PORCENTAJE AMBULATORIO"}

COVERAGE_BLOCK_PATTERN = re.compile(r"(\d+)%\s*(?:SIN TOPE)?", re.IGNORECASE)
LIBRE_ELECCION_PATTERN = re.compile(r"libre\s*elecci[oó]n", re.IGNORECASE)

COVERAGE_TEXT_FIXES: list[tuple[str, str]] = [
    ("Hospital\nClínico del Sur", "Hospital Clínico del Sur"),
    ("Hospital\nClinico del Sur", "Hospital Clínico del Sur"),
    ("Los\nAndes de Los Ángeles", "Los Andes de Los Ángeles"),
    ("Clínica Alemana\nde Osorno", "Clínica Alemana de Osorno"),
    ("Clínica Alemana \nde Osorno", "Clínica Alemana de Osorno"),
    ("UCChristus", "UC Christus"),
    ("Clínicas Puerto Montt, Biobío,", "Clínicas Puerto Montt, Clínica Biobío,"),
    ("Clínicas Puerto Montt, Biobío", "Clínicas Puerto Montt, Clínica Biobío"),
]

CLINIC_ALIASES: dict[str, str] = {
    "biobio": "cl-biobio",
    "clinica davila": "cl-davila",
    "clinica bupa santiago": "cl-bupa-santiago",
    "clinica bupa stgo": "cl-bupa-santiago",
    "clinica bupan santiago": "cl-bupa-santiago",
    "clinica davila vespucio": "cl-davila-vespucio",
    "clinica cordillera": "cl-cordillera",
    "clinica hospital del profesor": "cl-hospital-del-profesor",
    "clinica indisa": "cl-indisa-providencia-anexo",
    "clinica meds": "cl-meds",
    "clinica med": "cl-meds",
    "clinica santa maria": "cl-santa-maria",
    "hospital clinico uc christus clinica santa maria": "cl-santa-maria",
    "hospital clinico uc": "hosp-clinico-uc",
    "hospital clinico universidad catolica": "hosp-clinico-uc",
    "clinica uc": "hosp-clinico-uc",
    "clinica las condes": "cl-las-condes",
    "clinica san carlos de apoquindo": "cl-san-carlos",
    "clinica universidad de los andes": "cl-univ-andes",
    "clinica universidad de los andes": "cl-univ-andes",
    "integramedica": "integramedica",
    "isamedica": "integramedica",
    "clinica isamedica": "integramedica",
    "clinica tarapaca": "cl-tarapaca",
    "clinica san jose": "cl-san-jose-arica",
    "clinica la portada": "cl-la-portada",
    "clinica rcr atacama": "cl-atacama",
    "clinica atacama": "cl-atacama",
    "clinica bupa antofagasta": "cl-bupa-antofagasta",
    "clinica los leones": "cl-los-leones",
    "clinica los carrera": "cl-los-carrera",
    "hospital clinico vina del mar": "hosp-vina-del-mar",
    "clinica ciudad del mar": "cl-ciudad-del-mar",
    "clinica bupa renaca": "cl-bupa-renaca",
    "clinicas puerto montt": "cl-puerto-montt",
    "clinica puerto montt": "cl-puerto-montt",
    "clinica puerto montt achs salud": "cl-puerto-montt",
    "clinica biobio": "cl-biobio",
    "hospital clinico del sur": "hosp-clinico-del-sur",
    "clinica del sur achs salud": "hosp-clinico-del-sur",
    "lircay de talca": "cl-andes-salud-talca",
    "clinica lircay achs salud": "cl-andes-salud-talca",
    "los andes de los angeles": "cl-los-andes-la",
    "clinica los andes los angeles": "cl-los-andes-la",
    "clinica alemana de osorno": "cl-alemana-osorno",
    "clinica alemana de temuco": "cl-alemana-temuco",
    "clinica alemana de valdivia": "cl-alemana-valdivia",
    "clinica alemana valdivia": "cl-alemana-valdivia",
    "clinica alemana temuco": "cl-alemana-temuco",
    "clinica redsalud providencia": "cl-redsalud-providencia",
    "clinica redsalud santiago": "cl-redsalud-santiago",
    "clinica redsalud vitacura": "cl-redsalud-vitacura",
    "centros medicos redsalud": "cm-redsalud",
    "centros medicos red uc christus": "red-uc-christus",
    "hospital clinico universidad de chile": "hosp-clinico-uch",
    "clinica bupa santiago": "cl-bupa-santiago",
    "clinica redsalud mayor": "cl-redsalud-mayor",
    "clinica redsalud magallanes": "cl-redsalud-magallanes",
    "clinica andes salud puerto montt": "cl-andes-salud-puerto-montt",
    "clinica puerto varas": "cl-puerto-varas",
    "sanatorio aleman": "sanatorio-aleman",
    "hospital clinico uc christus": "red-uc-christus",
    "hospital clinico ucchristus": "red-uc-christus",
    "h clinico uc christus": "red-uc-christus",
    "red cm uc christus": "red-uc-christus",
    "red cm ucchristus": "red-uc-christus",
    "red uc christus": "red-uc-christus",
    "red ucchristus": "red-uc-christus",
    "red uc christus e": "red-uc-christus",
    "red cm uc christus e": "red-uc-christus",
    "clinica rcr atacama": "cl-atacama",
    "clinica u de los andes": "cl-univ-andes",
    "clinica uc san carlos": "cl-san-carlos",
    "clinica alemana valdivia": "cl-alemana-valdivia",
    "clinica alemana temuco": "cl-alemana-temuco",
    "clinica los andes de los angeles": "cl-los-andes-la",
}

CLINIC_SPLIT_PATTERN = re.compile(
    r"(?<=[.])\s+|,\s*|"
    r"(?=(?:Clínicas|Clínica|Clinica|Hospital|Centros|Red\s+CM|Red\s+UC|Integram|Isamédica|Lircay|Los Andes|Sanatorio)\b)",
    re.IGNORECASE,
)

AUTO_CLINIC_IDS: dict[str, str] = {}


def normalize_key(value: str) -> str:
    text = unicodedata.normalize("NFD", value)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.lower()
    text = text.replace("clinica", "clinica")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def slugify_clinic_id(name: str) -> str:
    key = normalize_key(name)
    if key in CLINIC_ALIASES:
        return CLINIC_ALIASES[key]
    if key in AUTO_CLINIC_IDS:
        return AUTO_CLINIC_IDS[key]

    slug = re.sub(r"[^a-z0-9]+", "-", key).strip("-")
    clinic_id = f"cl-{slug}" if slug else "cl-desconocida"
    AUTO_CLINIC_IDS[key] = clinic_id
    print(f"Clínica nueva (id autogenerado): {name!r} -> {clinic_id}", file=sys.stderr)
    return clinic_id


def resolve_clinic(name: str) -> tuple[str, str]:
    cleaned = name.strip().strip(",").strip(".")
    if not cleaned:
        return "", ""

    key = normalize_key(cleaned)
    clinic_id = CLINIC_ALIASES.get(key) or slugify_clinic_id(cleaned)
    return clinic_id, cleaned


def parse_price(value) -> float | None:
    if value is None:
        return None
    try:
        return round(float(str(value).strip().replace(",", ".")), 2)
    except ValueError:
        return None


def repair_coverage_text(text: str) -> str:
    for broken, fixed in COVERAGE_TEXT_FIXES:
        text = text.replace(broken, fixed)
    return repair_wrapped_lines(text)


def repair_wrapped_lines(text: str) -> str:
    """Une líneas partidas por saltos de línea dentro de un mismo bloque de cobertura."""
    lines = text.split("\n")
    merged: list[str] = []
    buffer = ""

    def flush() -> None:
        nonlocal buffer
        if buffer:
            merged.append(buffer)
            buffer = ""

    for line in lines:
        stripped = line.strip()
        if not stripped:
            flush()
            merged.append("")
            continue

        if COVERAGE_BLOCK_PATTERN.match(stripped):
            flush()
            merged.append(stripped)
            continue

        continuation = (
            buffer
            and (
                stripped.startswith("de ")
                or stripped.startswith("del ")
                or stripped.startswith("Clínico")
                or stripped.startswith("Clinico")
                or stripped.startswith("Andes ")
                or (not stripped[0].isupper() and stripped[0].isalpha())
                or buffer.endswith(",")
                or buffer.endswith("Hospital")
                or buffer.endswith("Los")
                or buffer.endswith("Clínica Alemana")
            )
        )

        if continuation:
            buffer = f"{buffer.rstrip(',')} {stripped.lstrip(',')}".strip()
        else:
            flush()
            buffer = stripped

    flush()
    return "\n".join(merged)


def split_clinic_token(token: str) -> list[str]:
    token = token.strip().strip(",").strip(".")
    if not token:
        return []

    parts = [part.strip().strip(",").strip(".") for part in CLINIC_SPLIT_PATTERN.split(token) if part.strip()]
    if len(parts) <= 1:
        return [token] if token else []

    return [part for part in parts if part]


def parse_coverage_text(text: str, coverage_type: str) -> list[dict]:
    if not text:
        return []

    text = repair_coverage_text(text)
    entries: list[dict] = []
    matches = list(COVERAGE_BLOCK_PATTERN.finditer(text))

    if not matches:
        return entries

    for index, match in enumerate(matches):
        percentage = int(match.group(1))
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        section = text[start:end].strip()

        if LIBRE_ELECCION_PATTERN.search(section):
            entries.append(
                {
                    "clinic_id": f"mv-libre-eleccion-{coverage_type[:1]}",
                    "clinic_name": "Libre Elección",
                    "percentage": percentage,
                    "type": coverage_type,
                }
            )
            continue

        for raw_line in section.split("\n"):
            for clinic_name in split_clinic_token(raw_line):
                if LIBRE_ELECCION_PATTERN.search(clinic_name):
                entries.append(
                    {
                        "clinic_id": f"mv-libre-eleccion-{coverage_type[:1]}",
                        "clinic_name": "Libre Elección",
                        "percentage": percentage,
                        "type": coverage_type,
                    }
                )
                continue

            clinic_id, resolved_name = resolve_clinic(clinic_name)
            if not clinic_id:
                continue

            entries.append(
                {
                    "clinic_id": clinic_id,
                    "clinic_name": resolved_name,
                    "percentage": percentage,
                    "type": coverage_type,
                }
            )

    return entries


def product_label_from_filename(path: Path) -> str:
    name = path.stem.replace("BASE DE DATO ", "").strip()
    return name


def parse_workbook(xlsx_path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    worksheet = workbook.active
    product_label = product_label_from_filename(xlsx_path)
    plans: list[dict] = []

    for row_idx in range(2, worksheet.max_row + 1):
        code_raw = worksheet.cell(row_idx, 1).value
        if code_raw is None or str(code_raw).strip() == "":
            continue

        code = str(code_raw).strip()
        if normalize_key(code) in {normalize_key(item) for item in PLAN_CODE_HEADERS}:
            continue

        price = parse_price(worksheet.cell(row_idx, 2).value)
        hosp_text = worksheet.cell(row_idx, 3).value
        amb_text = worksheet.cell(row_idx, 4).value

        coverage: list[dict] = []
        if hosp_text:
            coverage.extend(parse_coverage_text(str(hosp_text), "hospitalaria"))
        if amb_text:
            coverage.extend(parse_coverage_text(str(amb_text), "ambulatoria"))

        has_top = any(
            "100%" in str(hosp_text or "") and "sin tope" in str(hosp_text or "").lower()
            for _ in [0]
        )

        plans.append(
            {
                "isapre": ISAPRE_NAME,
                "plan_name": f"Nueva Masvida {product_label} {code}",
                "unique_code": code,
                "base_price_uf": price if price is not None else 0,
                "has_top": has_top,
                "additional_notes": product_label,
                "pdf_url": None,
                "pdf_public_id": None,
                "coverage": coverage,
            }
        )

    return plans


def parse_folder(folder: Path) -> list[dict]:
    all_plans: list[dict] = []
    seen_codes: set[str] = set()

    for xlsx_path in sorted(folder.glob("*.xlsx")):
        file_plans = parse_workbook(xlsx_path)
        print(f"{xlsx_path.name}: {len(file_plans)} planes", file=sys.stderr)

        for plan in file_plans:
            code = plan["unique_code"]
            if code in seen_codes:
                print(f"Código duplicado omitido: {code}", file=sys.stderr)
                continue
            seen_codes.add(code)
            all_plans.append(plan)

    return all_plans


def main() -> None:
    folder = Path(sys.argv[1]).expanduser().resolve() if len(sys.argv) > 1 else DEFAULT_FOLDER
    output = Path(sys.argv[2]).expanduser().resolve() if len(sys.argv) > 2 else OUTPUT_PATH

    if not folder.is_dir():
        print(f"Carpeta no encontrada: {folder}", file=sys.stderr)
        sys.exit(1)

    plans = parse_folder(folder)
    output.write_text(json.dumps(plans, ensure_ascii=False, indent=2), encoding="utf-8")

    with_cov = sum(1 for plan in plans if plan["coverage"])
    print(f"Total planes Masvida: {len(plans)} (con cobertura: {with_cov}) -> {output}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Genera Excel de estado Banmédica: planes creados, con PDF y faltantes."""

from __future__ import annotations

import json
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
OK_FILL = PatternFill("solid", fgColor="DCFCE7")
MISSING_FILL = PatternFill("solid", fgColor="FEE2E2")


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len("" if cell.value is None else str(cell.value)) for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 52)


def style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def main() -> None:
    if len(sys.argv) < 2:
        print("Uso: build-banmedica-status-xlsx.py <report.json> [salida.xlsx]", file=sys.stderr)
        sys.exit(1)

    report_path = Path(sys.argv[1]).expanduser().resolve()
    output_path = (
        Path(sys.argv[2]).expanduser().resolve()
        if len(sys.argv) > 2
        else report_path.parent / "banmedica-estado-planes-pdf.xlsx"
    )

    report = json.loads(report_path.read_text(encoding="utf-8"))
    all_plans = report["gaps"]["plansMissingPdf"] + [
        {
            "unique_code": item["unique_code"],
            "plan_name": item["plan_name"],
            "base_price_uf": item["base_price_uf"],
            "coverage_count": item["coverage_count"],
            "zones": item["zones"],
            "region": item["region"],
            "pdf_status": "Con PDF",
        }
        for item in json.loads((report_path.parent / ".banmedica-all-plans.json").read_text(encoding="utf-8"))
        if item.get("pdf_status") == "Con PDF"
    ]

    # Prefer authoritative all-plans export if present
    all_plans_path = report_path.parent / ".banmedica-all-plans.json"
    if all_plans_path.exists():
        all_plans = json.loads(all_plans_path.read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()

    # Resumen
    ws = wb.active
    ws.title = "Resumen"
    db = report["database"]
    storage = report["storage"]
    ws["A1"] = "Reporte Banmédica — planes y PDFs"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    summary_rows = [
        ("PDFs en storage (Santiago)", storage["pdfsSantiago"]),
        ("PDFs en storage (Regiones)", storage["pdfsRegiones"]),
        ("PDFs en storage (total)", storage["pdfsTotal"]),
        ("Planes en BD", db["total"]),
        ("Planes con PDF", db["withPdf"]),
        ("Planes sin PDF", db["withoutPdf"]),
        ("% con PDF", db["pdfPct"]),
        ("Planes con cobertura", db["withCoverage"]),
        ("Planes sin cobertura", db["withoutCoverage"]),
        ("PDF en storage sin plan en BD", len(report["gaps"]["pdfInStorageNotInDb"])),
    ]
    start = 4
    ws.cell(row=start, column=1, value="Indicador").font = Font(bold=True)
    ws.cell(row=start, column=2, value="Valor").font = Font(bold=True)
    for idx, (label, value) in enumerate(summary_rows, start=start + 1):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=value)

    autosize_columns(ws)

    # Todos los planes
    headers = [
        "Región",
        "Código único",
        "Nombre del plan",
        "Precio base (UF)",
        "Coberturas",
        "Zonas",
        "Estado PDF",
        "PDF en storage",
        "Archivo esperado",
    ]
    ws_all = wb.create_sheet("Todos los planes")
    ws_all.append(headers)
    style_header_row(ws_all, 1, len(headers))

    for plan in sorted(all_plans, key=lambda p: (p.get("region", ""), p["unique_code"])):
        pdf_status = plan.get("pdf_status") or ("Con PDF" if plan.get("pdf_url") else "Sin PDF")
        pdf_in_storage = plan.get("pdf_in_storage", "")
        row = [
            plan.get("region", ""),
            plan["unique_code"],
            plan["plan_name"],
            plan["base_price_uf"],
            plan.get("coverage_count", 0),
            plan.get("zones", ""),
            pdf_status,
            pdf_in_storage,
            f"{plan['unique_code']}.pdf",
        ]
        ws_all.append(row)
        current = ws_all.max_row
        fill = OK_FILL if pdf_status == "Con PDF" else MISSING_FILL
        ws_all.cell(row=current, column=7).fill = fill

    ws_all.freeze_panes = "A2"
    autosize_columns(ws_all)

    # Faltantes
    ws_missing = wb.create_sheet("Sin PDF")
    ws_missing.append(headers)
    style_header_row(ws_missing, 1, len(headers))
    missing = [p for p in all_plans if (p.get("pdf_status") or "") != "Con PDF"]
    for plan in sorted(missing, key=lambda p: (p.get("region", ""), p["unique_code"])):
        ws_missing.append(
            [
                plan.get("region", ""),
                plan["unique_code"],
                plan["plan_name"],
                plan["base_price_uf"],
                plan.get("coverage_count", 0),
                plan.get("zones", ""),
                "Sin PDF",
                plan.get("pdf_in_storage", "No"),
                f"{plan['unique_code']}.pdf",
            ]
        )
    ws_missing.freeze_panes = "A2"
    autosize_columns(ws_missing)

    # PDF orphan
    ws_orphan = wb.create_sheet("PDF sin plan")
    ws_orphan.append(["Código PDF", "Archivo"])
    style_header_row(ws_orphan, 1, 2)
    for code in report["gaps"]["pdfInStorageNotInDb"]:
        ws_orphan.append([code, f"{code}.pdf"])
    autosize_columns(ws_orphan)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel generado: {output_path}")


if __name__ == "__main__":
    main()

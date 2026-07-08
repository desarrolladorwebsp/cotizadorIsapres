#!/usr/bin/env python3
"""Genera Excel con catálogo de clínicas, zonas faltantes y definición de zonas."""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
OK_FILL = PatternFill("solid", fgColor="DCFCE7")
MISSING_FILL = PatternFill("solid", fgColor="FEE2E2")
WARN_FILL = PatternFill("solid", fgColor="FEF9C3")
ZONE_FILL = PatternFill("solid", fgColor="E0F2FE")


def autosize_columns(ws, max_width: int = 72) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len("" if cell.value is None else str(cell.value)) for cell in column_cells
        )
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), max_width)


def style_header_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_resumen_sheet(wb, payload: dict) -> None:
    ws = wb.active
    ws.title = "Resumen"
    summary = payload["summary"]

    ws["A1"] = "Reporte de clínicas y zonas geográficas"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"] = (
        f"Clínicas: {summary['total_clinicas']} | "
        f"Con zona: {summary['con_zona']} | "
        f"Sin zona: {summary['sin_zona']} | "
        f"Coberturas en catálogo: {summary['total_coberturas']}"
    )

    headers = ["Métrica", "Valor"]
    start = 5
    for col, header in enumerate(headers, start=1):
        ws.cell(row=start, column=col, value=header)
    style_header_row(ws, start, len(headers))

    rows = [
        ("Total clínicas / prestadores", summary["total_clinicas"]),
        ("Clínicas con zona asignada", summary["con_zona"]),
        ("Clínicas sin zona (requieren revisión)", summary["sin_zona"]),
        ("Zonas definidas en el cotizador", summary["zonas_definidas"]),
        ("Total entradas de cobertura", summary["total_coberturas"]),
    ]

    row_idx = start + 1
    for label, value in rows:
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if label.startswith("Clínicas sin zona") and value > 0:
            ws.cell(row=row_idx, column=2).fill = MISSING_FILL
        elif label.startswith("Clínicas con zona"):
            ws.cell(row=row_idx, column=2).fill = OK_FILL
        row_idx += 1

    ws.cell(row=row_idx + 1, column=1, value="Notas").font = Font(bold=True)
    ws.cell(
        row=row_idx + 2,
        column=1,
        value=(
            "Cada clínica debe tener al menos una zona. Las zonas se usan en el filtro "
            "«Filtrado por Zona» del cotizador. Libre elección se asigna a todas las zonas."
        ),
    )
    ws.merge_cells(start_row=row_idx + 2, start_column=1, end_row=row_idx + 2, end_column=2)
    ws.cell(row=row_idx + 2, column=1).alignment = Alignment(wrap_text=True)

    autosize_columns(ws)


def build_zone_definitions_sheet(wb, payload: dict) -> None:
    ws = wb.create_sheet("Definición de zonas")
    headers = [
        "ID zona",
        "Nombre",
        "Grupo",
        "Descripción",
        "Áreas / comunas",
        "Zona padre",
        "Ejemplos de prestadores",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    group_labels = {
        "rm_wide": "Región Metropolitana (amplia)",
        "rm_sector": "Sector RM",
        "region": "Región fuera de RM",
    }

    for zone in payload["zone_definitions"]:
        ws.append(
            [
                zone["id"],
                zone["label"],
                group_labels.get(zone["group"], zone["group"]),
                zone["description"],
                zone["areas"],
                zone["parent_zone_id"] or "—",
                zone["example_providers"],
            ]
        )

    for row in range(2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = ZONE_FILL

    ws.freeze_panes = "A2"
    autosize_columns(ws, max_width=80)


def build_clinic_sheet(wb, title: str, rows: list[dict], highlight_missing: bool) -> None:
    ws = wb.create_sheet(title)
    headers = [
        "ID clínica",
        "Nombre",
        "Estado",
        "Zonas (ID)",
        "Zonas (nombre)",
        "Planes vinculados",
        "Coberturas",
        "Isapres",
        "En tabla clínicas",
        "Notas",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for row in rows:
        ws.append(
            [
                row["clinic_id"],
                row["clinic_name"],
                row["estado"],
                ", ".join(row["zone_ids"]),
                row["zone_labels"],
                row["planes_vinculados"],
                row["coberturas"],
                row["isapres"],
                "Sí" if row["en_tabla_clinics"] else "No",
                row["notas"],
            ]
        )
        excel_row = ws.max_row
        if highlight_missing:
            ws.cell(row=excel_row, column=3).fill = MISSING_FILL
        elif row["estado"] == "Asignada":
            ws.cell(row=excel_row, column=3).fill = OK_FILL
        if not row["en_tabla_clinics"]:
            ws.cell(row=excel_row, column=9).fill = WARN_FILL

    ws.freeze_panes = "A2"
    autosize_columns(ws, max_width=96)


def main() -> None:
    if len(sys.argv) < 3:
        print(
            "Uso: build-clinic-zones-xlsx.py <report.json> <salida.xlsx>",
            file=sys.stderr,
        )
        sys.exit(1)

    input_path = Path(sys.argv[1]).expanduser().resolve()
    output_path = Path(sys.argv[2]).expanduser().resolve()
    payload = json.loads(input_path.read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    build_resumen_sheet(wb, payload)
    build_zone_definitions_sheet(wb, payload)

    missing = payload.get("missing", [])
    if missing:
        build_clinic_sheet(wb, "Sin zona", missing, highlight_missing=True)
    else:
        ws = wb.create_sheet("Sin zona")
        ws["A1"] = "Todas las clínicas tienen zona asignada."
        ws["A1"].font = Font(bold=True, color="166534")

    build_clinic_sheet(wb, "Catálogo completo", payload["clinics"], highlight_missing=False)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(
        f"Excel generado: {output_path} "
        f"({payload['summary']['total_clinicas']} clínicas, "
        f"{payload['summary']['sin_zona']} sin zona)"
    )


if __name__ == "__main__":
    main()

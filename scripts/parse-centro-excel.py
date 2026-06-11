#!/usr/bin/env python3
"""Parsea BASE DE DATOS CENTRO.xlsx a JSON para import-centro-plans.ts."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

CLINIC_NAME_TO_ID = {
    "Centros Médicos RedSalud(A.3)": "cm-redsalud",
    "Clínica Bupa Reñaca": "cl-bupa-renaca",
    "Clínica Ciudad del Mar": "cl-ciudad-del-mar",
    "Clínica Indisa Maipú": "cl-indisa-maipu",
    "Clínica Indisa Providencia": "cl-indisa-providencia-anexo",
    "Clínicam Indisa Providencia": "cl-indisa-providencia-anexo",
    "Clínica Las Condes": "cl-las-condes",
    "Clínica Las Condes(A.3)": "cl-las-condes",
    "Clínica Los Carrera": "cl-los-carrera",
    "Clínica Los Leones": "cl-los-leones",
    "Clínica Meds": "cl-meds",
    "Clínica Meds(A.3)": "cl-meds",
    "Clínica RedSalud Rancagua": "cl-redsalud-rancagua",
    "Clínica RedSalud Valparaiso": "cl-redsalud-valparaiso",
    "Clínica RedSalud Valparaíso": "cl-redsalud-valparaiso",
    "Clínica RedSalud Vitacura": "cl-redsalud-vitacura",
    "Clínica San Carlos de Apoquindo": "cl-san-carlos",
    "Clínica Santa Maria": "cl-santa-maria",
    "Clínica Santa María(A.3)": "cl-santa-maria",
    "Clínica Universidad de los Andes": "cl-univ-andes",
    "Hospital Clinico Fusat": "hosp-clinico-fusat",
    "Hospital Clinico de Viña del Mar": "hosp-vina-del-mar",
    "Hospital Clínico Universidad Católica": "hosp-clinico-uc",
    "Integramédica": "integramedica",
    "Red de Salud UC Christus": "red-uc-christus",
}

OUTPUT_PATH = Path(__file__).resolve().parent / ".centro-plans-parsed.json"


def parse_price(value) -> float | None:
    if value is None:
        return None
    return float(str(value).strip().replace(",", "."))


def parse_coverage_text(text: str, coverage_type: str) -> list[dict]:
    if not text:
        return []

    entries: list[dict] = []
    header_pattern = re.compile(r"(\d+)%\s+Sin Tope")
    matches = list(header_pattern.finditer(text))

    for index, match in enumerate(matches):
        percentage = int(match.group(1))
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        section = text[start:end]

        for line in section.split("\n"):
            clinic_name = line.strip()
            if not clinic_name:
                continue

            clinic_id = CLINIC_NAME_TO_ID.get(clinic_name)
            if not clinic_id:
                print(f"Clínica sin mapeo (omitida): {clinic_name!r}", file=sys.stderr)
                continue

            entries.append(
                {
                    "clinic_id": clinic_id,
                    "clinic_name": clinic_name,
                    "percentage": percentage,
                    "type": coverage_type,
                }
            )

    return entries


def parse_workbook(xlsx_path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    worksheet = workbook["Hoja 1"]
    plans: list[dict] = []

    for row_idx in range(1, worksheet.max_row + 1):
        if worksheet.cell(row_idx, 1).value != "CODIGO":
            continue

        codes: list[str] = []
        columns: list[int] = []
        for col in range(2, worksheet.max_column + 1):
            value = worksheet.cell(row_idx, col).value
            if value:
                codes.append(str(value).strip())
                columns.append(col)

        if not codes:
            continue

        prices: dict[str, float | None] = {}
        hospitalario: dict[str, str | None] = {}
        ambulatorio: dict[str, str | None] = {}

        for row in range(row_idx + 1, min(row_idx + 10, worksheet.max_row + 1)):
            label = worksheet.cell(row, 1).value
            if label == "CODIGO":
                break

            for code, col in zip(codes, columns):
                cell_value = worksheet.cell(row, col).value
                if label == "PRECIO BASE":
                    prices[code] = parse_price(cell_value)
                elif label == "PORCENTAJE HOSPITALARIO":
                    hospitalario[code] = cell_value
                elif label == "PORCENTAJE AMBULATORIO":
                    ambulatorio[code] = cell_value

        for code in codes:
            coverage: list[dict] = []
            hosp_text = hospitalario.get(code)
            amb_text = ambulatorio.get(code)

            if hosp_text:
                coverage.extend(parse_coverage_text(str(hosp_text), "hospitalaria"))
            if amb_text:
                coverage.extend(parse_coverage_text(str(amb_text), "ambulatoria"))

            plans.append(
                {
                    "isapre": "Consalud",
                    "plan_name": f"Consalud {code}",
                    "unique_code": code,
                    "base_price_uf": prices.get(code),
                    "has_top": False,
                    "additional_notes": None,
                    "pdf_url": None,
                    "pdf_public_id": None,
                    "coverage": coverage,
                }
            )

    return plans


def main() -> None:
    if len(sys.argv) < 2:
        print("Uso: parse-centro-excel.py <ruta-al-xlsx>", file=sys.stderr)
        sys.exit(1)

    xlsx_path = Path(sys.argv[1]).expanduser().resolve()
    plans = parse_workbook(xlsx_path)
    OUTPUT_PATH.write_text(json.dumps(plans, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Planes parseados: {len(plans)} -> {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

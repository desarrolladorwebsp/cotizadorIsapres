#!/usr/bin/env python3
"""Parsea BASE DE DATOS PLANES VIDA TRES.xlsx a JSON."""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

ISAPRE_NAME = "Vida Tres"
DEFAULT_XLSX = (
    Path(__file__).resolve().parent.parent
    / "storage"
    / "planes-pdf"
    / "vida tres"
    / "BASE DE DATOS PLANES VIDA TRES.xlsx"
)
OUTPUT_PATH = Path(__file__).resolve().parent / ".vidatres-plans-parsed.json"

COVERAGE_BLOCK_PATTERN = re.compile(r"(\d+)%\s*(?:SIN TOPE)?", re.IGNORECASE)
LIBRE_ELECCION_PATTERN = re.compile(r"libre\s*elecci[oó]n", re.IGNORECASE)

CLINIC_ALIASES: dict[str, str] = {
    "clinica davila": "cl-davila",
    "clinica davila vespucio": "cl-davila-vespucio",
    "clinica vespucio": "cl-davila-vespucio",
    "centros medicos red davila": "cm-red-davila",
    "clinica indisa": "cl-indisa-providencia-anexo",
    "clinica indisa providencia": "cl-indisa-providencia-anexo",
    "clinica santa maria": "cl-santa-maria",
    "centros medicos clinica santa maria": "cm-santa-maria",
    "hospital clinico uc": "hosp-clinico-uc",
    "clinica uc": "hosp-clinico-uc",
    "clinica redsalud santiago": "cl-redsalud-santiago",
    "clinica redsalud providencia": "cl-redsalud-providencia",
    "clinica redsalud valparaiso": "cl-redsalud-valparaiso",
    "clinica redsalud elqui": "cl-redsalud-elqui",
    "clinica redsalud iquique": "cl-redsalud-iquique",
    "clinica redsalud magallanes": "cl-redsalud-magallanes",
    "clinica redsalud mayor temuco": "cl-redsalud-mayor",
    "clinica san carlos de apoquindo": "cl-san-carlos",
    "clinica universidad de los andes": "cl-univ-andes",
    "integramedica": "integramedica",
    "clinica isamedica": "integramedica",
    "isamedica": "integramedica",
    "vidaintegra": "vidaintegra",
    "clinica las condes": "cl-las-condes",
    "clinica meds": "cl-meds",
    "clinica ciudad del mar": "cl-ciudad-del-mar",
    "clinica bupa renaca": "cl-bupa-renaca",
    "clinica los carrera interclinica": "cl-los-carrera",
    "clinica los carrera": "cl-los-carrera",
    "centros medicos red uc christus": "red-uc-christus",
    "centros medicos redsalud": "cm-redsalud",
    "hospital clinico fusat": "hosp-clinico-fusat",
    "hospital clinico vina del mar": "hosp-vina-del-mar",
    "clinica alemana de valdivia": "cl-alemana-valdivia",
    "clinica alemana de osorno": "cl-alemana-osorno",
    "clinica alemana": "cl-alemana-santiago-a3",
    "clinica andes salud puerto montt": "cl-andes-salud-puerto-montt",
    "clinica portada achs salud": "cl-la-portada",
    "clinica la portada": "cl-la-portada",
    "clinica puerto varas": "cl-puerto-varas",
    "clinica puerto montt achs salud": "cl-puerto-montt",
    "clinica san jose interclinica": "cl-san-jose-arica",
    "clinica bupa santiago": "cl-bupa-santiago",
    "clinica bupa antofagasta": "cl-bupa-antofagasta",
    "clinica atacama": "cl-atacama",
    "clinica tarapaca": "cl-tarapaca",
    "clinica biobio": "cl-biobio",
    "clinica alemana de temuco": "cl-alemana-temuco",
    "clinica los andes los angeles": "cl-los-andes-la",
    "clinica andes salud talca": "cl-andes-salud-talca",
    "clinica andes salud concepcion": "cl-andes-salud-concepcion",
    "clinica cordillera": "cl-cordillera",
    "clinica hospital del profesor": "cl-hospital-del-profesor",
    "hospital clinico universidad de chile": "hosp-clinico-uch",
    "sanatorio aleman": "sanatorio-aleman",
    "clinica tarapaca inter": "cl-tarapaca",
    "centromed": "cl-centromed",
}

CLINIC_SPLIT_PATTERN = re.compile(
    r"(?<=[.])\s+|,\s*|"
    r"(?=(?:Clínicas|Clínica|Clinica|Hospital|Centros|Red\s+CM|Red\s+UC|Integram|Isamédica|Vidaintegra|Lircay|Los Andes|Sanatorio)\b)",
    re.IGNORECASE,
)

AUTO_CLINIC_IDS: dict[str, str] = {}


def normalize_key(value: str) -> str:
    text = unicodedata.normalize("NFD", value)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_plan_code(raw) -> str | None:
    text = str(raw).strip()
    if not text or text.upper() == "CODIGO":
        return None

    match = re.match(r"^([A-Za-z0-9]+)", text)
    if not match:
        return None

    return match.group(1).upper()


SKIP_CLINIC_KEYS = {"no aplica", "n a", "na", ")", "("}


def is_invalid_clinic_name(name: str) -> bool:
    cleaned = name.strip().strip(",").strip(".")
    if not cleaned:
        return True
    if cleaned in {")", "(", "-", "–", "—"}:
        return True
    if len(cleaned) <= 2 and not cleaned.isalnum():
        return True
    return normalize_key(cleaned) in SKIP_CLINIC_KEYS


def slugify_clinic_id(name: str) -> str:
    key = normalize_key(name)
    if key in CLINIC_ALIASES:
        return CLINIC_ALIASES[key]
    if key in AUTO_CLINIC_IDS:
        return AUTO_CLINIC_IDS[key]

    slug = re.sub(r"[^a-z0-9]+", "-", key).strip("-")
    clinic_id = f"cl-{slug}" if slug else "cl-desconocida"
    AUTO_CLINIC_IDS[key] = clinic_id
    print(f"Clínica nueva (id autogenerado): {name!r} -> {clinic_id}", file=sys.stderr)
    return clinic_id


def resolve_clinic(name: str) -> tuple[str, str]:
    cleaned = name.strip().strip(",").strip(".")
    if is_invalid_clinic_name(cleaned):
        return "", ""

    key = normalize_key(cleaned)
    if key in SKIP_CLINIC_KEYS:
        return "", ""
    clinic_id = CLINIC_ALIASES.get(key) or slugify_clinic_id(cleaned)
    return clinic_id, cleaned


def parse_price(value) -> float | None:
    if value is None:
        return None
    try:
        return round(float(str(value).strip().replace(",", ".")), 2)
    except ValueError:
        return None


def split_clinic_token(token: str) -> list[str]:
    token = token.strip().strip(",").strip(".")
    if not token:
        return []

    parts = [
        part.strip().strip(",").strip(".")
        for part in CLINIC_SPLIT_PATTERN.split(token)
        if part.strip()
    ]
    return parts if parts else [token]


def parse_coverage_text(text: str, coverage_type: str) -> list[dict]:
    if not text:
        return []

    entries: list[dict] = []
    matches = list(COVERAGE_BLOCK_PATTERN.finditer(text))

    for index, match in enumerate(matches):
        percentage = int(match.group(1))
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        section = text[start:end].strip()

        if LIBRE_ELECCION_PATTERN.search(section):
            entries.append(
                {
                    "clinic_id": f"vt-libre-eleccion-{coverage_type[:1]}",
                    "clinic_name": "Libre Elección",
                    "percentage": percentage,
                    "type": coverage_type,
                }
            )
            continue

        for raw_line in section.split("\n"):
            for clinic_name in split_clinic_token(raw_line):
                if LIBRE_ELECCION_PATTERN.search(clinic_name):
                    entries.append(
                        {
                            "clinic_id": f"vt-libre-eleccion-{coverage_type[:1]}",
                            "clinic_name": "Libre Elección",
                            "percentage": percentage,
                            "type": coverage_type,
                        }
                    )
                    continue

                clinic_id, resolved_name = resolve_clinic(clinic_name)
                if not clinic_id:
                    continue

                entries.append(
                    {
                        "clinic_id": clinic_id,
                        "clinic_name": resolved_name,
                        "percentage": percentage,
                        "type": coverage_type,
                    }
                )

    return entries


def parse_workbook(xlsx_path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    worksheet = workbook.active
    plans: list[dict] = []
    seen_codes: set[str] = set()

    for row_idx in range(2, worksheet.max_row + 1):
        code = normalize_plan_code(worksheet.cell(row_idx, 1).value)
        if not code or code in seen_codes or code == "NONE":
            continue

        seen_codes.add(code)
        price = parse_price(worksheet.cell(row_idx, 2).value)
        hosp_text = worksheet.cell(row_idx, 3).value
        amb_text = worksheet.cell(row_idx, 4).value

        coverage: list[dict] = []
        if hosp_text:
            coverage.extend(parse_coverage_text(str(hosp_text), "hospitalaria"))
        if amb_text:
            coverage.extend(parse_coverage_text(str(amb_text), "ambulatoria"))

        has_top = bool(
            hosp_text
            and re.search(r"100\s*%", str(hosp_text))
            and "sin tope" not in str(hosp_text).lower()
        )

        plans.append(
            {
                "isapre": ISAPRE_NAME,
                "plan_name": f"Vida Tres {code}",
                "unique_code": code,
                "base_price_uf": price if price is not None else 0,
                "has_top": has_top,
                "additional_notes": None,
                "pdf_url": None,
                "pdf_public_id": None,
                "coverage": coverage,
            }
        )

    return plans


def main() -> None:
    xlsx_path = Path(sys.argv[1]).expanduser().resolve() if len(sys.argv) > 1 else DEFAULT_XLSX
    output = Path(sys.argv[2]).expanduser().resolve() if len(sys.argv) > 2 else OUTPUT_PATH

    if not xlsx_path.is_file():
        print(f"Archivo no encontrado: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    plans = parse_workbook(xlsx_path)
    output.write_text(json.dumps(plans, ensure_ascii=False, indent=2), encoding="utf-8")

    with_cov = sum(1 for plan in plans if plan["coverage"])
    print(f"Planes Vida Tres: {len(plans)} (con cobertura: {with_cov}) -> {output}")


if __name__ == "__main__":
    main()

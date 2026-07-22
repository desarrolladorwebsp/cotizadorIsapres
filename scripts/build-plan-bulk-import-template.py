#!/usr/bin/env python3
"""
Genera la plantilla Excel para carga / actualización masiva de planes.

Los campos restringidos son SELECT reales (Data Validation tipo lista):
- Listas embebidas en la MISMA hoja (columnas ocultas) → máxima compatibilidad
- errorStyle=stop → Excel rechaza valores fuera de la lista
- Catálogo visible Catalogo_Clinicas + hoja _Listas (respaldo)
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
CLINICS_PATH = ROOT / "src" / "assets" / "clinics.json"
OUT_DIR = ROOT / "docs" / "templates"
OUT_FILE = OUT_DIR / "plantilla-carga-masiva-planes.xlsx"

ISAPRES = [
    "Consalud",
    "Banmédica",
    "Colmena",
    "Cruz Blanca",
    "Vida Tres",
    "Nueva Masvida",
    "Esencial",
]

PLAN_TYPES = ["Preferente", "Libre Elección", "Cerrado"]
ACCIONES = ["crear", "actualizar", "eliminar"]
COVERAGE_TYPES = ["hospitalaria", "ambulatoria"]
PERCENTAGES = ["40", "50", "60", "70", "80", "90", "100"]

ZONES = [
    ("rm-metropolitana", "Región Metropolitana"),
    ("rm-norte", "RM Norte"),
    ("rm-sur", "RM Sur"),
    ("rm-oriente", "RM Oriente"),
    ("rm-poniente", "RM Poniente"),
    ("rm-centro", "RM Centro"),
    ("norte", "Zona Norte"),
    ("octava", "Octava Región"),
    ("valparaiso", "Valparaíso"),
    ("biobio", "Biobío"),
]

MAX_DATA_ROWS = 500
ZONE_COLUMNS = 5

# Columnas ocultas en cada hoja para alimentar los SELECT (misma hoja = más fiable)
PLANES_LIST_START_COL = 30  # AD
COBERTURAS_LIST_START_COL = 20  # T

HEADER_FILL = PatternFill("solid", fgColor="092558")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SELECT_FILL = PatternFill("solid", fgColor="0D6DEE")
CALC_FILL = PatternFill("solid", fgColor="E2E8F0")
CALC_HEADER_FILL = PatternFill("solid", fgColor="475569")
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
ERROR_FILL = PatternFill("solid", fgColor="FECACA")
TITLE_FONT = Font(bold=True, size=14, color="092558")
SUBTITLE_FONT = Font(size=11, color="334155")
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)


def load_clinics() -> list[tuple[str, str]]:
    raw = json.loads(CLINICS_PATH.read_text(encoding="utf-8"))
    clinics = [(str(c["id"]).strip(), str(c["name"]).strip()) for c in raw]
    clinics.sort(key=lambda item: item[1].casefold())
    return clinics


def zone_labels() -> list[str]:
    return [f"{zone_id} | {label}" for zone_id, label in ZONES]


def clinic_labels(clinics: list[tuple[str, str]]) -> list[str]:
    return [f"{clinic_id} | {name}" for clinic_id, name in clinics]


def style_header(
    ws,
    headers: list[str],
    select_headers: set[str],
    calc_headers: set[str] | None = None,
) -> None:
    calc_headers = calc_headers or set()
    for col, header in enumerate(headers, start=1):
        if header in select_headers:
            label = f"{header} ▾"
            fill = SELECT_FILL
        elif header in calc_headers:
            label = f"{header} 🔒"
            fill = CALC_HEADER_FILL
        else:
            label = header
            fill = HEADER_FILL
        cell = ws.cell(1, col, label)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN
        cell.protection = LOCKED
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{MAX_DATA_ROWS + 1}"


def write_hidden_list(
    ws,
    start_col: int,
    title: str,
    values: list[str],
) -> str:
    """Escribe una lista en columna oculta y devuelve el rango absoluto para DataValidation."""
    header = ws.cell(1, start_col, f"__{title}")
    header.font = Font(bold=True, color="94A3B8", size=9)

    for i, value in enumerate(values, start=2):
        ws.cell(i, start_col, value)

    end_row = 1 + len(values)
    col_letter = get_column_letter(start_col)
    ws.column_dimensions[col_letter].hidden = True
    # Rango same-sheet (sin nombre de hoja): más compatible con Excel/LibreOffice
    return f"${col_letter}$2:${col_letter}${end_row}"


def add_select(
    ws,
    list_range: str,
    cells: str,
    *,
    allow_blank: bool,
    prompt: str,
    error: str,
) -> None:
    """SELECT estricto: solo valores de la lista; Excel bloquea cualquier otro."""
    dv = DataValidation(
        type="list",
        formula1=list_range,
        allow_blank=allow_blank,
        showDropDown=False,  # False = mostrar flecha del desplegable
        showErrorMessage=True,
        showInputMessage=True,
        errorStyle="stop",
        errorTitle="Selección obligatoria",
        error=error,
        promptTitle="Selecciona de la lista",
        prompt=prompt,
    )
    ws.add_data_validation(dv)
    dv.add(cells)


def build_listas_sheet(wb: openpyxl.Workbook, clinics: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet("_Listas")

    catalogs: list[tuple[str, list[str]]] = [
        ("accion", ACCIONES),
        ("isapre", ISAPRES),
        ("tipo_plan", PLAN_TYPES),
        ("zona", zone_labels()),
        ("tipo_cobertura", COVERAGE_TYPES),
        ("porcentaje", PERCENTAGES),
        ("clinica", clinic_labels(clinics)),
    ]

    for col, (title, values) in enumerate(catalogs, start=1):
        ws.cell(1, col, title)
        for row, value in enumerate(values, start=2):
            ws.cell(row, col, value)

    # También id/nombre de clínica por separado
    ws.cell(1, 8, "clinica_id")
    ws.cell(1, 9, "clinica_nombre")
    for row, (clinic_id, name) in enumerate(clinics, start=2):
        ws.cell(row, 8, clinic_id)
        ws.cell(row, 9, name)

    named = {
        "ListaAcciones": f"'_Listas'!$A$2:$A${1 + len(ACCIONES)}",
        "ListaIsapres": f"'_Listas'!$B$2:$B${1 + len(ISAPRES)}",
        "ListaTiposPlan": f"'_Listas'!$C$2:$C${1 + len(PLAN_TYPES)}",
        "ListaZonas": f"'_Listas'!$D$2:$D${1 + len(ZONES)}",
        "ListaTipoCobertura": f"'_Listas'!$E$2:$E${1 + len(COVERAGE_TYPES)}",
        "ListaPorcentajes": f"'_Listas'!$F$2:$F${1 + len(PERCENTAGES)}",
        "ListaClinicas": f"'_Listas'!$G$2:$G${1 + len(clinics)}",
    }
    for name, attr in named.items():
        wb.defined_names.add(DefinedName(name=name, attr_text=attr))

    ws.sheet_state = "hidden"


def build_instrucciones(wb: openpyxl.Workbook, clinic_count: int) -> None:
    ws = wb.create_sheet("Instrucciones", 0)
    ws["A1"] = "Plantilla de carga / actualización masiva de planes"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    lines = [
        "",
        "CAMPOS SELECT (flecha ▾ — NO escribir a mano)",
        "Planes: accion, isapre, tipo_plan, zona_1…zona_5",
        "Coberturas: tipo_cobertura, clinica, porcentaje",
        "Excel BLOQUEA cualquier valor que no esté en la lista (error stop).",
        "",
        "COLUMNAS CALCULADAS (candado 🔒 — NO editables)",
        "alerta_codigo          Avisa si codigo_unico está duplicado en Planes.",
        "coberturas_hosp        Cuenta filas hospitalaria en Coberturas con el mismo codigo_unico.",
        "coberturas_amb         Cuenta filas ambulatoria en Coberturas con el mismo codigo_unico.",
        "Se actualizan solas al agregar/quitar filas en Coberturas. La hoja Planes está protegida.",
        "",
        "CÓMO USAR",
        "1. En cada celda SELECT haz clic y elige de la flecha desplegable.",
        "2. Completa Planes (1 fila = 1 plan) y Coberturas (N filas por codigo_unico).",
        "3. Revisa coberturas_hosp / coberturas_amb: deben subir al ir añadiendo clínicas.",
        "4. No inventes isapres, tipos, zonas ni clínicas.",
        "5. codigo_unico debe ser único dentro de este Excel.",
        "",
        "HOJA PLANES",
        "accion ▾*             crear | actualizar | eliminar",
        "codigo_unico*         texto libre (clave del plan)",
        "alerta_codigo 🔒      automática",
        "isapre ▾*             lista oficial",
        "nombre_plan*          texto",
        "precio_base_uf*       número ≥ 0",
        "ges_uf                número ≥ 0 (opcional)",
        "tipo_plan ▾*          Preferente | Libre Elección | Cerrado",
        "coberturas_hosp 🔒    # prestadores hospitalarios (desde Coberturas)",
        "coberturas_amb 🔒     # prestadores ambulatorios (desde Coberturas)",
        "zona_1…zona_5 ▾      una o varias (deja vacías las que no uses)",
        "notas                 texto libre",
        "",
        "HOJA COBERTURAS",
        "codigo_unico*         mismo código que en Planes",
        "tipo_cobertura ▾*     hospitalaria | ambulatoria",
        "clinica ▾*            id | nombre (catálogo embebido)",
        "porcentaje ▾*         40 | 50 | 60 | 70 | 80 | 90 | 100",
        "",
        "REGLAS",
        f"- {clinic_count} clínicas en el SELECT (hoja Catalogo_Clinicas para consultar).",
        "- Si hay filas de Coberturas para un plan → al importar se REEMPLAZAN todas sus coberturas.",
        "- eliminar: basta accion + codigo_unico.",
        "- Encabezado azul brillante = SELECT. Gris con 🔒 = fórmula bloqueada. Navy = texto/número.",
        "",
        "EJEMPLO",
        "Filas EJEMPLO-001 / EJEMPLO-002 son de muestra: bórralas antes de una carga real.",
        "EJEMPLO-001 debería mostrar coberturas_hosp=2 y coberturas_amb=2.",
    ]

    for row, text in enumerate(lines, start=2):
        ws.cell(row, 1, text)
        if text.startswith(
            ("CAMPOS", "COLUMNAS", "CÓMO", "HOJA", "REGLAS", "EJEMPLO")
        ):
            ws.cell(row, 1).font = Font(bold=True, color="092558", size=12)
        elif text.startswith("-") or "▾" in text:
            ws.cell(row, 1).font = SUBTITLE_FONT

    ws.column_dimensions["A"].width = 110


def coverage_count_formula(row_idx: int, coverage_type: str) -> str:
    """Cuenta filas en Coberturas para el codigo_unico de esta fila y un tipo."""
    end = MAX_DATA_ROWS + 1
    return (
        f'=IF(B{row_idx}="","",'
        f'COUNTIFS(Coberturas!$A$2:$A${end},B{row_idx},'
        f'Coberturas!$B$2:$B${end},"{coverage_type}"))'
    )


def duplicate_code_formula(row_idx: int) -> str:
    end = MAX_DATA_ROWS + 1
    return (
        f'=IF(B{row_idx}="","",IF(COUNTIF($B$2:$B${end},B{row_idx})>1,'
        f'"DUPLICADO — revisa codigo_unico",""))'
    )


def build_planes_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Planes", 1)

    zone_headers = [f"zona_{i}" for i in range(1, ZONE_COLUMNS + 1)]
    headers = [
        "accion",
        "codigo_unico",
        "alerta_codigo",
        "isapre",
        "nombre_plan",
        "precio_base_uf",
        "ges_uf",
        "tipo_plan",
        "coberturas_hosp",
        "coberturas_amb",
        *zone_headers,
        "notas",
    ]
    select_headers = {"accion", "isapre", "tipo_plan", *zone_headers}
    calc_headers = {"alerta_codigo", "coberturas_hosp", "coberturas_amb"}
    style_header(ws, headers, select_headers, calc_headers)

    # Índices 1-based
    col_by_header = {h: i + 1 for i, h in enumerate(headers)}
    editable_headers = {
        "accion",
        "codigo_unico",
        "isapre",
        "nombre_plan",
        "precio_base_uf",
        "ges_uf",
        "tipo_plan",
        *zone_headers,
        "notas",
    }

    # Listas ocultas en la misma hoja (fuente de los SELECT)
    col = PLANES_LIST_START_COL
    range_accion = write_hidden_list(ws, col, "accion", ACCIONES)
    col += 1
    range_isapre = write_hidden_list(ws, col, "isapre", ISAPRES)
    col += 1
    range_tipo = write_hidden_list(ws, col, "tipo_plan", PLAN_TYPES)
    col += 1
    range_zona = write_hidden_list(ws, col, "zona", zone_labels())

    samples = [
        {
            "accion": "crear",
            "codigo_unico": "EJEMPLO-001",
            "isapre": "Colmena",
            "nombre_plan": "PLAN EJEMPLO CREAR",
            "precio_base_uf": 3.81,
            "ges_uf": 0.778,
            "tipo_plan": "Preferente",
            "zona_1": "rm-metropolitana | Región Metropolitana",
            "zona_2": "valparaiso | Valparaíso",
            "notas": "Fila de ejemplo — reemplazar",
        },
        {
            "accion": "actualizar",
            "codigo_unico": "EJEMPLO-002",
            "isapre": "Banmédica",
            "nombre_plan": "PLAN EJEMPLO ACTUALIZAR",
            "precio_base_uf": 2.5,
            "tipo_plan": "Cerrado",
            "zona_1": "rm-oriente | RM Oriente",
        },
    ]

    end = MAX_DATA_ROWS + 1

    def apply_calc_and_protection(row_idx: int) -> None:
        for header, col_idx in col_by_header.items():
            cell = ws.cell(row_idx, col_idx)
            cell.border = THIN
            if header == "alerta_codigo":
                cell.value = duplicate_code_formula(row_idx)
                cell.fill = WARN_FILL
                cell.protection = LOCKED
            elif header == "coberturas_hosp":
                cell.value = coverage_count_formula(row_idx, "hospitalaria")
                cell.fill = CALC_FILL
                cell.protection = LOCKED
                cell.alignment = Alignment(horizontal="center")
            elif header == "coberturas_amb":
                cell.value = coverage_count_formula(row_idx, "ambulatoria")
                cell.fill = CALC_FILL
                cell.protection = LOCKED
                cell.alignment = Alignment(horizontal="center")
            elif header in editable_headers:
                cell.protection = UNLOCKED
            else:
                cell.protection = LOCKED

    for row_idx, sample in enumerate(samples, start=2):
        for header, col_idx in col_by_header.items():
            if header in calc_headers:
                continue
            value = sample.get(header, "")
            ws.cell(row_idx, col_idx, value if value != "" else None)
        apply_calc_and_protection(row_idx)

    for row_idx in range(4, end + 1):
        apply_calc_and_protection(row_idx)

    # SELECT estrictos (misma hoja)
    add_select(
        ws,
        range_accion,
        f"A2:A{end}",
        allow_blank=False,
        prompt="Elige: crear, actualizar o eliminar.",
        error="Solo puedes seleccionar: crear, actualizar o eliminar.",
    )
    add_select(
        ws,
        range_isapre,
        f"D2:D{end}",
        allow_blank=False,
        prompt="Elige una isapre de la lista oficial.",
        error="Isapre no válida. Selecciona una opción del desplegable.",
    )
    add_select(
        ws,
        range_tipo,
        f"H2:H{end}",
        allow_blank=False,
        prompt="Elige Preferente, Libre Elección o Cerrado.",
        error="Tipo de plan inválido. Solo hay 3 opciones en el desplegable.",
    )
    zona_start = col_by_header["zona_1"]
    for offset in range(ZONE_COLUMNS):
        letter = get_column_letter(zona_start + offset)
        add_select(
            ws,
            range_zona,
            f"{letter}2:{letter}{end}",
            allow_blank=True,
            prompt="Elige una zona (o deja vacío).",
            error="Zona no válida. Selecciona del desplegable o deja la celda vacía.",
        )

    dv_price = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Número inválido",
        error="Ingresa un número ≥ 0 (UF).",
    )
    ws.add_data_validation(dv_price)
    dv_price.add(f"F2:F{end}")
    dv_price.add(f"G2:G{end}")

    ws.conditional_formatting.add(
        f"B2:B{end}",
        FormulaRule(
            formula=[f'AND(B2<>"",COUNTIF($B$2:$B${end},B2)>1)'],
            fill=ERROR_FILL,
        ),
    )
    ws.conditional_formatting.add(
        f"C2:C{end}",
        FormulaRule(
            formula=['C2="DUPLICADO — revisa codigo_unico"'],
            fill=ERROR_FILL,
            font=Font(bold=True, color="991B1B"),
        ),
    )

    widths = {
        "A": 14,
        "B": 16,
        "C": 30,
        "D": 16,
        "E": 36,
        "F": 14,
        "G": 10,
        "H": 18,
        "I": 14,
        "J": 14,
        "K": 36,
        "L": 36,
        "M": 36,
        "N": 36,
        "O": 36,
        "P": 28,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    ws.cell(
        MAX_DATA_ROWS + 3,
        1,
        "▾ = SELECT. 🔒 = fórmula bloqueada (coberturas_hosp / coberturas_amb cuentan "
        "filas de Coberturas por codigo_unico). Hoja protegida: no edites columnas grises.",
    ).font = Font(italic=True, color="64748B", size=9)
    ws.cell(MAX_DATA_ROWS + 3, 1).protection = LOCKED

    # Proteger hoja: columnas calculadas no editables; el resto sí.
    ws.protection.sheet = True
    ws.protection.enable()
    ws.protection.autoFilter = True
    ws.protection.sort = True
    ws.protection.insertRows = True
    ws.protection.deleteRows = False
    # Sin contraseña: se puede desproteger si hace falta, pero por defecto bloquea 🔒.


def build_coberturas_sheet(
    wb: openpyxl.Workbook, clinics: list[tuple[str, str]]
) -> None:
    ws = wb.create_sheet("Coberturas", 2)
    headers = ["codigo_unico", "tipo_cobertura", "clinica", "porcentaje"]
    select_headers = {"tipo_cobertura", "clinica", "porcentaje"}
    style_header(ws, headers, select_headers)

    col = COBERTURAS_LIST_START_COL
    range_tipo = write_hidden_list(ws, col, "tipo_cobertura", COVERAGE_TYPES)
    col += 1
    range_clinica = write_hidden_list(ws, col, "clinica", clinic_labels(clinics))
    col += 1
    range_pct = write_hidden_list(ws, col, "porcentaje", PERCENTAGES)

    by_id = {cid: name for cid, name in clinics}

    def resolve_clinic(token: str) -> str:
        clinic_id = token.split(" | ", 1)[0].strip()
        if clinic_id in by_id:
            return f"{clinic_id} | {by_id[clinic_id]}"
        cid, name = clinics[0]
        return f"{cid} | {name}"

    samples = [
        (
            "EJEMPLO-001",
            "hospitalaria",
            "cl-alemana-santiago-a | Clínica Alemana de Santiago(A.)",
            "100",
        ),
        ("EJEMPLO-001", "hospitalaria", "cl-las-condes | Clínica Las Condes", "100"),
        (
            "EJEMPLO-001",
            "ambulatoria",
            "cl-alemana-santiago-a | Clínica Alemana de Santiago(A.)",
            "40",
        ),
        ("EJEMPLO-001", "ambulatoria", "cm-redsalud | Centros Médicos RedSalud", "40"),
        ("EJEMPLO-002", "hospitalaria", "cl-davila | Clínica Dávila", "40"),
        ("EJEMPLO-002", "ambulatoria", "cl-davila | Clínica Dávila", "40"),
    ]

    for row_idx, (code, tipo, clinic, pct) in enumerate(samples, start=2):
        ws.cell(row_idx, 1, code).border = THIN
        ws.cell(row_idx, 2, tipo).border = THIN
        ws.cell(row_idx, 3, resolve_clinic(clinic)).border = THIN
        ws.cell(row_idx, 4, pct).border = THIN

    end = MAX_DATA_ROWS + 1
    add_select(
        ws,
        range_tipo,
        f"B2:B{end}",
        allow_blank=False,
        prompt="Elige hospitalaria o ambulatoria.",
        error="Solo hospitalaria o ambulatoria. Usa el desplegable.",
    )
    add_select(
        ws,
        range_clinica,
        f"C2:C{end}",
        allow_blank=False,
        prompt="Elige una clínica del catálogo (id | nombre).",
        error="Clínica no válida. Debes seleccionarla del desplegable.",
    )
    add_select(
        ws,
        range_pct,
        f"D2:D{end}",
        allow_blank=False,
        prompt="Elige 40, 50, 60, 70, 80, 90 o 100.",
        error="Porcentaje no permitido. Selecciona del desplegable.",
    )

    for letter, width in {"A": 16, "B": 18, "C": 52, "D": 14}.items():
        ws.column_dimensions[letter].width = width

    ws.cell(
        MAX_DATA_ROWS + 3,
        1,
        "▾ = SELECT. tipo_cobertura, clinica y porcentaje solo admiten valores de la lista.",
    ).font = Font(italic=True, color="64748B", size=9)


def build_catalogo_clinicas_visible(
    wb: openpyxl.Workbook, clinics: list[tuple[str, str]]
) -> None:
    ws = wb.create_sheet("Catalogo_Clinicas", 3)
    ws["A1"] = "id"
    ws["B1"] = "nombre"
    ws["C1"] = "valor_del_select"
    for col in range(1, 4):
        ws.cell(1, col).font = HEADER_FONT
        ws.cell(1, col).fill = HEADER_FILL
        ws.cell(1, col).border = THIN

    for i, (clinic_id, name) in enumerate(clinics, start=2):
        ws.cell(i, 1, clinic_id).border = THIN
        ws.cell(i, 2, name).border = THIN
        ws.cell(i, 3, f"{clinic_id} | {name}").border = THIN

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 52
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{len(clinics) + 1}"
    ws.cell(
        len(clinics) + 3,
        1,
        "Solo consulta. En Coberturas usa el SELECT de la columna clinica ▾.",
    ).font = Font(italic=True, color="64748B", size=9)


def main() -> None:
    clinics = load_clinics()
    if not clinics:
        raise SystemExit(f"No se encontraron clínicas en {CLINICS_PATH}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_listas_sheet(wb, clinics)
    build_instrucciones(wb, len(clinics))
    build_planes_sheet(wb)
    build_coberturas_sheet(wb, clinics)
    build_catalogo_clinicas_visible(wb, clinics)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)
    print(f"Plantilla regenerada: {OUT_FILE}")
    print("SELECT Planes: accion, isapre, tipo_plan, zona_1..5")
    print("CALC Planes (bloqueadas): alerta_codigo, coberturas_hosp, coberturas_amb")
    print("SELECT Coberturas: tipo_cobertura, clinica, porcentaje")
    print(f"Clínicas: {len(clinics)}")


if __name__ == "__main__":
    main()

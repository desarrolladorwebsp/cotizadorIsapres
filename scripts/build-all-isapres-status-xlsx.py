#!/usr/bin/env python3
"""Genera Excel con el estado de planes/PDF de todas las Isapres."""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
OK_FILL = PatternFill("solid", fgColor="DCFCE7")
WARN_FILL = PatternFill("solid", fgColor="FEF9C3")
MISSING_FILL = PatternFill("solid", fgColor="FEE2E2")

HEADERS = [
    "Isapre",
    "Total planes",
    "Con PDF",
    "Sin PDF",
    "% con PDF",
    "Con cobertura",
    "Sin cobertura",
    "Estado",
    "Códigos sin PDF",
]


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len("" if cell.value is None else str(cell.value)) for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 64)


def status_label(total: int, sin_pdf: int) -> str:
    if total == 0:
        return "Sin planes"
    if sin_pdf == 0:
        return "Completo"
    if sin_pdf < total:
        return "Incompleto"
    return "Sin PDFs"


def main() -> None:
    if len(sys.argv) < 3:
        print("Uso: build-all-isapres-status-xlsx.py <entrada.json> <salida.xlsx>", file=sys.stderr)
        sys.exit(1)

    input_path = Path(sys.argv[1]).expanduser().resolve()
    output_path = Path(sys.argv[2]).expanduser().resolve()
    rows = json.loads(input_path.read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Todas las Isapres"

    ws["A1"] = "Estado de planes por Isapre"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    total_planes = sum(row["total_planes"] for row in rows)
    total_pdf = sum(row["con_pdf"] for row in rows)
    total_missing = sum(row["sin_pdf"] for row in rows)
    ws["A3"] = f"Total general: {total_planes} planes | {total_pdf} con PDF | {total_missing} sin PDF"

    start = 5
    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=start, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        estado = status_label(row["total_planes"], row["sin_pdf"])
        ws.append(
            [
                row["isapre"],
                row["total_planes"],
                row["con_pdf"],
                row["sin_pdf"],
                row["pct_pdf"],
                row["con_cobertura"],
                row["sin_cobertura"],
                estado,
                row["codigos_sin_pdf"],
            ]
        )
        current = ws.max_row
        if row["total_planes"] == 0:
            fill = WARN_FILL
        elif row["sin_pdf"] == 0 and row["total_planes"] > 0:
            fill = OK_FILL
        elif row["sin_pdf"] > 0:
            fill = MISSING_FILL
        else:
            fill = None
        if fill:
            ws.cell(row=current, column=8).fill = fill

    ws.freeze_panes = f"A{start + 1}"
    autosize_columns(ws)

    ws_codes = wb.create_sheet("Códigos sin PDF")
    ws_codes.append(["Isapre", "Cantidad", "Códigos"])
    for col in range(1, 4):
        cell = ws_codes.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row in rows:
        if not row["codigos_sin_pdf"]:
            continue
        codes = [code.strip() for code in row["codigos_sin_pdf"].split(",") if code.strip()]
        ws_codes.append([row["isapre"], len(codes), ", ".join(codes)])

    ws_codes.freeze_panes = "A2"
    autosize_columns(ws_codes)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel generado: {output_path} ({len(rows)} isapres)")


if __name__ == "__main__":
    main()

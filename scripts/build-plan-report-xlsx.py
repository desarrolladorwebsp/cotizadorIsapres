#!/usr/bin/env python3
"""Genera un único Excel con el reporte completo de planes por Isapre."""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
OK_FILL = PatternFill("solid", fgColor="DCFCE7")
WARN_FILL = PatternFill("solid", fgColor="FEF9C3")
MISSING_FILL = PatternFill("solid", fgColor="FEE2E2")
ISAPRE_FILL = PatternFill("solid", fgColor="E5E7EB")
ISAPRE_FONT = Font(bold=True)

DETAIL_HEADERS = [
    "Isapre",
    "ID plan",
    "Nombre del plan",
    "Precio base (UF)",
    "Coberturas",
    "Zonas",
    "Incluye TOP",
    "PDF esperado",
]


def autosize_columns(ws, max_width: int = 64) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len("" if cell.value is None else str(cell.value)) for cell in column_cells
        )
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), max_width)


def style_header_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def status_label(total: int, sin_pdf: int) -> str:
    if total == 0:
        return "Sin planes"
    if sin_pdf == 0:
        return "Completo"
    if sin_pdf < total:
        return "Incompleto"
    return "Sin PDFs"


def build_resumen_sheet(wb, summary: list[dict], missing: list[dict]) -> None:
    ws = wb.active
    ws.title = "Resumen"

    total_planes = sum(row["total_planes"] for row in summary)
    total_pdf = sum(row["con_pdf"] for row in summary)
    total_missing = len(missing)

    ws["A1"] = "Reporte de planes por Isapre"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"] = (
        f"Total catálogo: {total_planes} planes | "
        f"{total_pdf} con PDF | {total_missing} sin PDF | "
        f"{round(total_pdf / total_planes * 100, 1) if total_planes else 0}% con PDF"
    )

    headers = [
        "Isapre",
        "Total planes",
        "Con PDF",
        "Sin PDF",
        "% con PDF",
        "Con cobertura",
        "Sin cobertura",
        "Estado",
    ]
    start = 5
    for col, header in enumerate(headers, start=1):
        ws.cell(row=start, column=col, value=header)
    style_header_row(ws, start, len(headers))

    row_idx = start + 1
    for item in summary:
        estado = status_label(item["total_planes"], item["sin_pdf"])
        ws.cell(row=row_idx, column=1, value=item["isapre"])
        ws.cell(row=row_idx, column=2, value=item["total_planes"])
        ws.cell(row=row_idx, column=3, value=item["con_pdf"])
        ws.cell(row=row_idx, column=4, value=item["sin_pdf"])
        ws.cell(row=row_idx, column=5, value=item["pct_pdf"])
        ws.cell(row=row_idx, column=6, value=item["con_cobertura"])
        ws.cell(row=row_idx, column=7, value=item["sin_cobertura"])
        ws.cell(row=row_idx, column=8, value=estado)

        if item["total_planes"] == 0:
            ws.cell(row=row_idx, column=8).fill = WARN_FILL
        elif item["sin_pdf"] == 0:
            ws.cell(row=row_idx, column=8).fill = OK_FILL
        else:
            ws.cell(row=row_idx, column=8).fill = MISSING_FILL
        row_idx += 1

    ws.cell(row=row_idx + 1, column=1, value="TOTAL").font = Font(bold=True)
    for col, value in enumerate(
        [total_planes, total_pdf, total_missing, round(total_pdf / total_planes * 100, 1) if total_planes else 0],
        start=2,
    ):
        ws.cell(row=row_idx + 1, column=col, value=value).font = Font(bold=True)

    ws.freeze_panes = f"A{start + 1}"
    autosize_columns(ws)


def build_codigos_sheet(wb, summary: list[dict]) -> None:
    ws = wb.create_sheet("Códigos sin PDF")
    ws.append(["Isapre", "Cantidad", "Códigos sin PDF"])
    style_header_row(ws, 1, 3)

    for item in summary:
        if not item["codigos_sin_pdf"]:
            continue
        codes = [code.strip() for code in item["codigos_sin_pdf"].split(",") if code.strip()]
        ws.append([item["isapre"], len(codes), ", ".join(codes)])

    ws.freeze_panes = "A2"
    autosize_columns(ws, max_width=120)


def build_detalle_sheet(wb, missing: list[dict]) -> None:
    ws = wb.create_sheet("Detalle sin PDF")
    ws.append(DETAIL_HEADERS)
    style_header_row(ws, 1, len(DETAIL_HEADERS))

    current_isapre = None
    for row in missing:
        if row["isapre"] != current_isapre:
            current_isapre = row["isapre"]
            ws.append([])
            summary_row = ws.max_row + 1
            ws.cell(row=summary_row, column=1, value=current_isapre)
            ws.cell(row=summary_row, column=1).fill = ISAPRE_FILL
            ws.cell(row=summary_row, column=1).font = ISAPRE_FONT
            for col in range(2, len(DETAIL_HEADERS) + 1):
                ws.cell(row=summary_row, column=col).fill = ISAPRE_FILL

        ws.append(
            [
                row["isapre"],
                row["unique_code"],
                row["plan_name"],
                row["base_price_uf"],
                row["coverage_count"],
                row["zones"],
                "Sí" if row["has_top"] else "No",
                f"{row['unique_code']}.pdf",
            ]
        )

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def build_isapre_sheets(wb, missing: list[dict]) -> None:
    by_isapre: dict[str, list[dict]] = {}
    for row in missing:
        by_isapre.setdefault(row["isapre"], []).append(row)

    for isapre, rows in sorted(by_isapre.items(), key=lambda item: item[0].lower()):
        if not rows:
            continue
        safe_title = f"Faltantes {isapre}"[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(safe_title)
        ws.append(DETAIL_HEADERS)
        style_header_row(ws, 1, len(DETAIL_HEADERS))

        for row in rows:
            ws.append(
                [
                    row["isapre"],
                    row["unique_code"],
                    row["plan_name"],
                    row["base_price_uf"],
                    row["coverage_count"],
                    row["zones"],
                    "Sí" if row["has_top"] else "No",
                    f"{row['unique_code']}.pdf",
                ]
            )

        ws.freeze_panes = "A2"
        autosize_columns(ws)


def main() -> None:
    if len(sys.argv) < 4:
        print(
            "Uso: build-plan-report-xlsx.py <status.json> <missing.json> <salida.xlsx>",
            file=sys.stderr,
        )
        sys.exit(1)

    status_path = Path(sys.argv[1]).expanduser().resolve()
    missing_path = Path(sys.argv[2]).expanduser().resolve()
    output_path = Path(sys.argv[3]).expanduser().resolve()

    summary = json.loads(status_path.read_text(encoding="utf-8"))
    missing = json.loads(missing_path.read_text(encoding="utf-8"))
    missing.sort(key=lambda row: (row["isapre"].lower(), row["unique_code"]))

    wb = openpyxl.Workbook()
    build_resumen_sheet(wb, summary, missing)
    build_codigos_sheet(wb, summary)
    build_detalle_sheet(wb, missing)
    build_isapre_sheets(wb, missing)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel generado: {output_path} ({len(summary)} isapres, {len(missing)} sin PDF)")


if __name__ == "__main__":
    main()

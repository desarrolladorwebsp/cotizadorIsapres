#!/usr/bin/env python3
"""Genera un Excel de planes sin PDF a partir de JSON exportado desde la BD."""

from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "Isapre",
    "Código único",
    "Nombre del plan",
    "Precio base (UF)",
    "Coberturas",
    "Zonas",
    "Incluye TOP",
    "PDF esperado",
]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ISAPRE_FILL = PatternFill("solid", fgColor="E5E7EB")
ISAPRE_FONT = Font(bold=True)


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 64)


def load_all_isapres_summary(path: Path | None) -> list[dict]:
    if not path or not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def build_summary_sheet(wb, rows: list[dict], all_isapres: list[dict]) -> None:
    ws = wb.active
    ws.title = "Resumen"

    counts = Counter(row["isapre"] for row in rows)
    total_missing = len(rows)

    ws["A1"] = "Reporte de planes sin PDF"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"] = f"Total planes sin PDF: {total_missing}"

    headers = [
        "Isapre",
        "Total planes",
        "Con PDF",
        "Sin PDF",
        "% sin PDF",
        "Estado",
    ]
    start = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    row_idx = start + 1
    source = all_isapres if all_isapres else [
        {"isapre": isapre, "total_planes": count, "con_pdf": 0, "sin_pdf": count}
        for isapre, count in sorted(counts.items(), key=lambda item: item[0].lower())
    ]

    for item in source:
        total = item.get("total_planes", 0)
        sin_pdf = item.get("sin_pdf", counts.get(item["isapre"], 0))
        con_pdf = item.get("con_pdf", max(total - sin_pdf, 0))
        pct = round((sin_pdf / total) * 100, 1) if total else 0
        estado = "Completo" if total > 0 and sin_pdf == 0 else ("Sin planes" if total == 0 else "Faltan PDFs")

        ws.cell(row=row_idx, column=1, value=item["isapre"])
        ws.cell(row=row_idx, column=2, value=total)
        ws.cell(row=row_idx, column=3, value=con_pdf)
        ws.cell(row=row_idx, column=4, value=sin_pdf)
        ws.cell(row=row_idx, column=5, value=pct)
        ws.cell(row=row_idx, column=6, value=estado)
        row_idx += 1

    ws.cell(row=row_idx + 1, column=1, value="TOTAL SIN PDF").font = Font(bold=True)
    ws.cell(row=row_idx + 1, column=4, value=total_missing).font = Font(bold=True)

    autosize_columns(ws)


def build_codes_sheet(wb, rows: list[dict], all_isapres: list[dict]) -> None:
    ws = wb.create_sheet("Códigos por Isapre")
    ws.append(["Isapre", "Cantidad", "Códigos sin PDF"])
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    by_isapre: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        by_isapre[row["isapre"]].append(row["unique_code"])

    if all_isapres:
        ordered = [item["isapre"] for item in all_isapres]
    else:
        ordered = sorted(by_isapre.keys(), key=str.lower)

    seen = set()
    for isapre in ordered:
        seen.add(isapre)
        codes = sorted(by_isapre.get(isapre, []))
        ws.append([isapre, len(codes), ", ".join(codes) if codes else "—"])

    for isapre in sorted(set(by_isapre.keys()) - seen, key=str.lower):
        codes = sorted(by_isapre[isapre])
        ws.append([isapre, len(codes), ", ".join(codes)])

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def build_detail_sheet(wb, rows: list[dict]) -> None:
    ws = wb.create_sheet("Detalle por Isapre")

    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_isapre = None
    for row in rows:
        if row["isapre"] != current_isapre:
            current_isapre = row["isapre"]
            ws.append([])
            summary_row = ws.max_row + 1
            ws.cell(row=summary_row, column=1, value=current_isapre)
            ws.cell(row=summary_row, column=1).fill = ISAPRE_FILL
            ws.cell(row=summary_row, column=1).font = ISAPRE_FONT
            for col in range(2, len(HEADERS) + 1):
                ws.cell(row=summary_row, column=col).fill = ISAPRE_FILL

        ws.append(
            [
                row["isapre"],
                row["unique_code"],
                row["plan_name"],
                row["base_price_uf"],
                row["coverage_count"],
                row["zones"],
                "Sí" if row["has_top"] else "No",
                f"{row['unique_code']}.pdf",
            ]
        )

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def build_isapre_sheets(wb, rows: list[dict]) -> None:
    by_isapre: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        by_isapre[row["isapre"]].append(row)

    for isapre, isapre_rows in sorted(by_isapre.items(), key=lambda item: item[0].lower()):
        if not isapre_rows:
            continue
        safe_title = isapre[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(safe_title)
        ws.append(HEADERS)
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for row in isapre_rows:
            ws.append(
                [
                    row["isapre"],
                    row["unique_code"],
                    row["plan_name"],
                    row["base_price_uf"],
                    row["coverage_count"],
                    row["zones"],
                    "Sí" if row["has_top"] else "No",
                    f"{row['unique_code']}.pdf",
                ]
            )

        ws.freeze_panes = "A2"
        autosize_columns(ws)


def main() -> None:
    if len(sys.argv) < 3:
        print(
            "Uso: build-missing-pdfs-xlsx.py <entrada.json> <salida.xlsx> [all-isapres.json]",
            file=sys.stderr,
        )
        sys.exit(1)

    input_path = Path(sys.argv[1]).expanduser().resolve()
    output_path = Path(sys.argv[2]).expanduser().resolve()
    all_isapres_path = Path(sys.argv[3]).expanduser().resolve() if len(sys.argv) > 3 else None

    rows = json.loads(input_path.read_text(encoding="utf-8"))
    rows.sort(key=lambda row: (row["isapre"].lower(), row["unique_code"]))
    all_isapres = load_all_isapres_summary(all_isapres_path)

    wb = openpyxl.Workbook()
    build_summary_sheet(wb, rows, all_isapres)
    build_codes_sheet(wb, rows, all_isapres)
    build_detail_sheet(wb, rows)
    build_isapre_sheets(wb, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel generado: {output_path} ({len(rows)} filas)")


if __name__ == "__main__":
    main()
